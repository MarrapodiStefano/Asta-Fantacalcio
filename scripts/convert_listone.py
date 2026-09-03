from openpyxl import load_workbook
from pathlib import Path
import json, re, hashlib, unicodedata

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Listone.xlsx"
TARGET = ROOT / "players.js"

def norm(value):
    s = str(value or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z0-9%]+", "", s)
    return s

ALIASES = {
    "id": {"id", "codice", "codicegiocatore", "playerid"},
    "role": {"r", "ruolo", "role"},
    "name": {"nome", "giocatore", "name"},
    "team": {"squadra", "team", "club"},
    "credits": {"quotazione", "quota", "qt", "qta", "crediti", "credits"},
    "pct": {"%", "percentuale", "pct", "percentualeacquisto", "percentualecrediti"},
    "pmv": {"pmv", "prezzomediovendita", "prezzomedio"},
    "appeal": {"appetibilita", "appeal"},
}

def find_column(headers, keys):
    for i, h in enumerate(headers):
        if h in keys:
            return i
    return None

wb = load_workbook(SOURCE, data_only=True, read_only=True)
ws = wb.active

rows = ws.iter_rows(values_only=True)
header_row = next(rows)
headers = [norm(v) for v in header_row]

cols = {key: find_column(headers, aliases) for key, aliases in ALIASES.items()}
required = ["role", "name", "team"]
missing = [k for k in required if cols[k] is None]
if missing:
    raise SystemExit(
        "Colonne obbligatorie non trovate: " + ", ".join(missing) +
        ". Intestazioni trovate: " + ", ".join(str(v) for v in header_row)
    )

players = []
seen = set()

for row in rows:
    if not any(v is not None and str(v).strip() for v in row):
        continue

    def value(key, default=None):
        i = cols[key]
        return row[i] if i is not None and i < len(row) else default

    role = str(value("role", "")).strip().upper()
    name = str(value("name", "")).strip()
    team = str(value("team", "")).strip()

    if role not in {"P", "D", "C", "A"} or not name:
        continue

    raw_id = value("id")
    if raw_id is None or str(raw_id).strip() == "":
        raw_id = int(hashlib.sha1(f"{role}|{name}|{team}".encode("utf-8")).hexdigest()[:8], 16)

    try:
        pid = int(raw_id)
    except Exception:
        pid = int(hashlib.sha1(str(raw_id).encode("utf-8")).hexdigest()[:8], 16)

    if pid in seen:
        pid = int(hashlib.sha1(f"{pid}|{role}|{name}|{team}".encode("utf-8")).hexdigest()[:8], 16)
    seen.add(pid)

    def number(key):
        x = value(key)
        if x is None or str(x).strip() == "":
            return None
        if isinstance(x, str):
            x = x.replace(",", ".").replace("%", "")
        try:
            n = float(x)
            return int(n) if n.is_integer() else n
        except Exception:
            return None

    players.append({
        "id": pid,
        "role": role,
        "name": name,
        "team": team,
        "credits": number("credits"),
        "pct": number("pct"),
        "pmv": number("pmv"),
        "appeal": number("appeal"),
    })

order = {"P": 0, "D": 1, "C": 2, "A": 3}
players.sort(key=lambda p: (order.get(p["role"], 9), -(p["credits"] or 0), p["name"].lower()))

TARGET.write_text(
    "const PLAYERS=" + json.dumps(players, ensure_ascii=False, separators=(",", ":")) + ";\n",
    encoding="utf-8"
)

print(f"Generati {len(players)} giocatori in {TARGET.name}")
